#!/usr/bin/env python
#coding=UTF-8

import sys
sys.path.append('/home/analise_liWGSseqs/software/biomart')
sys.path.append("/usr/lib/pymodules/python2.7/openpyxl")
from biomart import BiomartServer
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
import time
import cgi
import cgitb
import itertools
import test_best_hit
from collections import OrderedDict
cgitb.enable()

def indexx():
	#print 'Content-type: text/html; charset=utf-8'
	print """
 <html>
 <meta name="viewport" content="width=device-width, initial-scale=1.0">
<link href="https://fonts.googleapis.com/css?family=Arial" rel="stylesheet">
<style>
* {
  box-sizing: border-box;
}

body {
  background-color: #f1f1f1;
}

#regForm {
  background-color: #ffffff;
  margin: 80px auto;
  font-family: Arial;
  padding: 40px;
  width: 70%;
  min-width: 300px;
}

hgh {
  text-align: center;
  font-family: Arial;
}

aaa {
  padding:40px;
  font-family: Arial;
}

h1 {
  text-align: center;
  font-family: Arial;
}

head2 {
  padding: 10px;
  width: 100%;
  font-size: 17px;
  font-family: Arial;
  border: 1px solid #aaaaaa;
}

input {
  padding: 10px;
  width: 100%;
  font-size: 17px;
  font-family: Arial;
  border: 1px solid #aaaaaa;
}

/* Mark input boxes that gets an error on validation: */
input.invalid {
  background-color: #ffdddd;
}

/* Hide all steps by default: */
.tab {
  display: block
}

.nn {
  display: none;
}

button {
  font-family: Arial;
}

button:hover {
  opacity: 0.8;
}

#prevBtn {
  background-color: #bbbbbb;
}

/* Make circles that indicate the steps of the form: */
.step {
  height: 15px;
  width: 15px;
  margin: 0 2px;
  background-color: #bbbbbb;
  border: none;  
  border-radius: 50%;
  display: inline-block;
  opacity: 0.5;
}

.step.active {
  opacity: 1;
}

/* Mark the steps that are finished and valid: */
.step.finish {
  background-color: #4CAF50;
}

input[type=checkbox]
{
  /* Double-sized Checkboxes */
  -ms-transform: scale(1.2); /* IE */
  -moz-transform: scale(1.2); /* FF */
  -webkit-transform: scale(1.2); /* Safari and Chrome */
  -o-transform: scale(2); /* Opera */
  padding: 10px;
}

</style>

</html>
        
"""

#indexx()


def search(filt, att, version):
        """makes the search against the biomart database, using as filter arguments
        filt and retriving the atributes at att"""
        if version=="hg19":
                server = BiomartServer( "http://feb2014.archive.ensembl.org/biomart/" )
                hg=server.datasets["hsapiens_gene_ensembl"]
        elif version=="hg38":
                server=BiomartServer( "http://jul2018.archive.ensembl.org/biomart/" )
                hg=server.datasets["hsapiens_gene_ensembl"]
        response = hg.search({
          'filters':filt,
          'attributes': att
        })
        return response

def calc_sz(val, start, end):
	if val.endswith("%"):
		pp=float(val[:-1])/100
		sz=int(pp*(int(end)-int(start)))
	elif val.endswith("kb"):
		sz=int(val[:-2])*1000
	elif val.endswith("bp"):
		sz=int(val[:-2])
	else:
		sz=int(val)
	return sz

def parse_search(response, bps, val):###################################################
	disrupted=set()
	dis_cord={}
	deleted=set()
	for line in response.iter_lines():
		e=line.split("\t")
		if e[-1]=="protein_coding" or e[-1].endswith("lincRNA"):
			qqq=calc_sz(val, e[0],e[1])
			d=[str(int(e[0])-qqq),str(int(e[1])+qqq), e[2]]
			if (int(d[0])<=int(bps[0]) and int(d[1])>= int(bps[0])) or (int(d[0])<=int(bps[1]) and int(d[1])>=int(bps[1])):
				disrupted.add(d[2])
				dis_cord[d[2]]=[d[0],d[1]]
			if (int(d[0])>=int(bps[0]) and int(d[1])<=int(bps[1])):
				deleted.add(d[2])
				dis_cord[d[2]]=[d[0],d[1]]
	return '; '.join(disrupted), '; '.join(deleted),dis_cord

def select_cytoband(infile, ch, start,end):
	f=open(infile)
	band={}
	for i in f:
		line=i.split("\t")
		b=""
		if "chr"+ch==line[0]:
			if int(start)>=int(line[1]) and int(end)<=int(line[2]):
				return ch+line[-1].strip()
			if int(start)>=int(line[1]) and int(start)<=int(line[2]):
				b+=line[-1].strip()
			if int(end)>=int(line[1]) and int(end)<=int(line[2]):
				b+=line[-1].strip()
				return ch+b
	f.close()
	return band

def interpret_array_nomenclature(l):
	#oo=l[4:].replace(" ","")
	#aa=oo.replace("]", ";")
	bb=l.replace("p",";")
	cc=bb.replace("q", ";")
	dd=cc.split(";")
	ee=dd[-1].split("(")
	ff=ee[1].split(")")
	gg=ff[0].replace(",","")
	hh=gg.split("-")
	return dd[0], hh[0], hh[1]

def make_array_nomenclature(ch, start, end, version):
	if version=="hg38":
		v="GRCh38"
		band="/usr/local/apache/cgi-bin/bp2omim/cytoband_hg38.bed"
	if version=="hg19":
		v="GRCh37"
		band="/usr/local/apache/cgi-bin/bp2omim/cytoband_hg19.bed"
	b=select_cytoband(band, ch, start,end)
	name="["+v+"]"+b+"("+start+"-"+end+")"
	return name


def define_bed(n, chrr, start,end, start2, end2, version, chrrs, dic, vale, is_large):#start2 e end e sem 25kb
	dis_cord={}
	disrupted=""
	deleted=""
	if n=="":
		name=make_array_nomenclature(chrr, start, end, version)
		filt1= {'chromosome_name': chrr, 'start': [start2], 'end':[end2]}
		if is_large==False:
			disrupted, deleted, dis_cord=parse_search(search(filt1,at1,version), [start2,end2], vale)
		dic[name]=[chrr,[int(start)], [int(end)], disrupted, deleted]
		chrrs.add(chrr)
		if len(dis_cord)>0:
			for key, val in dis_cord.items():
				dic[key]=[chrr, [int(val[0])], [int(val[1])], "",""]
	else:
		filt1= {'chromosome_name': chrr, 'start': [str(start2)], 'end':[str(end2)]}
		if is_large==False:
			disrupted, deleted, dis_cord=parse_search(search(filt1,at1,version),[start2, end2], vale)
		dic[n]=[chrr, [int(start)], [int(end)], disrupted, deleted]
		chrrs.add(chrr)
		if len(dis_cord)>0:
			for key, val in dis_cord.items():
				dic[key]=[chrr, [int(val[0])], [int(val[1])], "",""]
	return dic, chrrs


def read_bed(bedfile, version, val):#le lista de coordenadas em bed
	"""Output dic - dic[name_type-array]=[chr, [start], [end], str(disrupted), str(deleted)]"""
	chrrs=set()
	dic=OrderedDict()
	for i in bedfile.split("\n"):
		if len(i)>5:
			if "q" in i or "p" in i:#com nomenclatura tipo array
				l=i.strip()
				ch,st,en=interpret_array_nomenclature(i)
				c=[ch]
				cords=[str(st),str(en)]
			else:#quando e dado como UCSC coordinates
				l=i.strip()
				aa=i.split(":")
				c=[aa[0].replace("chr","")]
				cords=aa[1].split("-")
			dic, chrrs=define_bed(l,c[0], cords[0].replace(",",""), cords[1].replace(",",""),cords[0].replace(",",""), cords[1].replace(",",""), version, chrrs, dic, val, False)
	return dic,chrrs

def read_bps(bedfile, version, val, flanki):#le lista de coordenadas em bed
	"""Output dic - dic[name_type-array]=[chr, [start], [end], str(disrupted), str(deleted)]"""
	flank=int(flanki)
	chrrs=set()
	dic=OrderedDict()
	for i in bedfile.split("\n"):
		if len(i)>5:
			if "_" in i:
				l=i.strip()
				cc=i.split("_")[1]
				dd=cc.split(":")
				c=[dd[0].replace("chr","")]
				if len(dd[1].split("-"))>1:
					cords=dd[1].split("-")
				else:
					cords=[dd[1], str(int(dd[1])+1)]
			else:
				l=i.strip()
				aa=i.split(":")
				c=[aa[0].replace("chr","")]
				cords=aa[1].split("-")
			dic, chrrs=define_bed(l,c[0], cords[0].replace(",",""), cords[1].replace(",",""),cords[0].replace(",",""), cords[1].replace(",",""), version, chrrs, dic, val, False)#normal
			if flank>0:
				dic, chrrs=define_bed(l+"_flank",c[0], str(int(cords[0].replace(",",""))-flank), str(int(cords[1].replace(",",""))+flank),str(int(cords[0].replace(",",""))-flank), str(int(cords[1].replace(",",""))+flank), version, chrrs, dic, 0, True)#com flanking
	return dic,chrrs



def read_bd(infile, chrrs):#todos excepto coe
	f=open(infile)
	dic_ch={}#dic[ch]=[start, stop, name]
	dic_name={}#dic[name]=[type, support]
	for i in f:
		line=i.split("\t")
		if i.startswith("#"):
			date=i.strip("#")
		elif len(line)>1:
			if line[1] in chrrs:
				if line[1] not in dic_ch:
					dic_ch[line[1]]=[[int(line[2]), int(line[3]), line[0]]]
				else:
					dic_ch[line[1]].append([int(line[2]), int(line[3]), line[0]])
				aa=[]
				for ff,b in itertools.izip(line[6].split(";"),line[8].split(";")):
					aa.append('=HYPERLINK("https://www.ncbi.nlm.nih.gov/pubmed/'+b+'","'+b+'")')
				if len(aa)==1:
					aa.append("")
				dic_name[line[0]]=['=HYPERLINK("'+line[7]+'","'+line[0]+'")', line[4], line[5]]
				dic_name[line[0]]+=aa
				dic_name[line[0]].append(line[-1].strip())
	f.close()
	return dic_ch, dic_name, date

def read_bd_coe(infile, chrrs):#so coe#primeiro genes depois nomenclatira
	f=open(infile)
	dic_ch={}#dic[ch]=[start, stop, name]
	dic_name={}#dic[name]=[type, support]
	for i in f:
		line=i.split("\t")
		if i.startswith("#"):
			date=i.strip("#")
		elif line[1] in chrrs:
			if line[1] not in dic_ch:
				dic_ch[line[1]]=[[int(line[2]), int(line[3]), line[0]]]
			else:
				dic_ch[line[1]].append([int(line[2]), int(line[3]), line[0]])
			aa=[]
			for ff,b in itertools.izip(line[6].split(";"),line[8].split(";")):
				aa.append('=HYPERLINK("https://www.ncbi.nlm.nih.gov/pubmed/'+b+'","'+b+'")')
			if len(aa)==1:
				aa.append("")
			dic_name[line[0]]=['=HYPERLINK("'+line[7]+'","'+line[0]+'")', line[4], line[5]]
			dic_name[line[0]]+=aa
			if len(line)==10:
				dic_name[line[0]].append(line[-1].strip())
			if len(line)==11:
				dic_name[line[0]].append(line[-1].strip())
				dic_name[line[0]].append(line[-2].strip())
	f.close()
	return dic_ch, dic_name, date
	
def get_Overlap(a,b):
	return max(0,min(a[1],b[1])-max(a[0],b[0]))

def get_ranges(perc):
	reanges=[[1,10],[10,20],[20,30],[30,40],[40,50],[50,60],[60,70],[70,80],[80,90],[90,100]]
	a=0
	di=[]
	while a<len(reanges):
		if reanges[a][0]>=int(perc):
			g=reanges[a:]
			break
		a+=1
	for el in g:
		di.append([])
	return g, di	

def complete_stats(di, ranges, perc_del, nome, freq, ref, typee):
	a=0
	while a<len(ranges):
		if perc_del >= ranges[a][0] and perc_del<=ranges[a][1]:
			di[a].append(nome)
			di[a].append(perc_del)
			di[a].append(freq)
			di[a].append(ref)
			di[a].append(typee)
		a+=1
	return di


def cross_values(clusters, bd_ch, bd_name, perc, stats, is_bp, ovl):#stats por bds, excepto clinVar/Gen #, is_bp
	v=OrderedDict()#tamanho_delecao; nome das alteracoes; cobertura; coberto/tamanho da delecao; tipo; suporte
	brek_ranges=[[1,10],[10,20],[20,30],[30,40],[40,50],[50,60],[60,70],[70,80],[80,90],[90,100]]
	dii=[[],[],[],[],[],[],[],[],[],[]]
	ranges, di=get_ranges(perc)
	variants=[]
	at=[]
	in_output=set()
	best_overlap=[]#list of the best overlap with [name of the best overlap, frequency of the best overlap, overlap mercentage]
	best_freq=[]#list of the best frequency with[name of the best frequency, frequency, overlap percentage]
	best_hits={}#dictionary with the best hits as: best_hits[key]=[name_best_overlap, best_overlap, name
	for key,value in clusters.items():
		ch=value[0]
		start=max(value[1])
		end=min(value[2])
		if key not in stats:
			ranges, di=get_ranges(perc)
			if is_bp==False:
				temp=di
			else:
				ranges=brek_ranges
				temp=[[],[],[],[],[],[],[],[],[],[]]
		if key in stats:
			temp=stats[key]
		if ch in bd_ch:
			for el in bd_ch[ch]:
				if el[0]!=el[1]:
					cov=get_Overlap((start,end),(el[0], el[1]))
					perc_del=round((float(cov)/(float(end-start)))*100.0, 3)
					perc_anot=round((float(cov)/(float(el[1]-el[0])))*100.0, 3)
					if (cov>0 and perc_del>=float(perc) and perc_anot>=float(perc) and ovl=="mutual") or (cov>0 and perc_del>=float(perc) and perc_anot>=float(perc) and str(start) not in key) or (is_bp==True and cov>0 and "_" in key) or (cov>0 and str(start) in key and perc_del==100.0 and ovl=="full" and is_bp==False):
						temp=complete_stats(temp, ranges, perc_del, bd_name[el[-1]][-1], bd_name[el[-1]][2].split(" ")[0],bd_name[el[-1]][3], bd_name[el[-1]][1])
						tm=[]
						in_output.add(key)
						at.append(key)
						at.append(str(end-start))
						at.append(bd_name[el[-1]][0])
						at.append(str(el[1]-el[0]))
						at.append(bd_name[el[-1]][-1])
						at.append(bd_name[el[-1]][1])
						at.append(bd_name[el[-1]][2])
						if "HYPERLINK" in bd_name[el[-1]][-2]:
							at.append(bd_name[el[-1]][-2])
						if "HYPERLINK" in bd_name[el[-1]][-3]:
							at.append(bd_name[el[-1]][-3])
						if "HYPERLINK" not in bd_name[el[-1]][-3] or "HYPERLINK" not in bd_name[el[-1]][-2]:
							at.append("")
						at.append(str(perc_del))
						at.append(str(perc_anot))
						variants.append(at)
						at=[]
		stats[key]=temp
		temp=[]
		v[key]=variants
		variants=[]
	rr=check_if_in_output(in_output, clusters, v)
	return rr, stats, ranges

def complement_stats(lg, flag):#4 em 4
	a=0
	tot=0#total of samples
	sup=0#suporting samples
	ref=set()#set of references
	ran=[]#list with the overlap to calculate the range
	names=""#string with the alterations
	hits=str(len(lg)/5)
	typee=set()
	while(a<len(lg)):
		names+=lg[a]+";"
		ran.append(float(lg[a+1]))##################################
		v=lg[a+2].split("/")
		typee.add(lg[a+4])
		if lg[a+3] in ref:
			sup+=float(v[0])
		else:
			ref.add(lg[a+3])
			sup+=float(v[0])
			tot+=float(v[1])
		a+=5
	r,f=calc_range_and_freq(ran, sup, tot)
	return [hits, ", ".join(typee), r, f]

def calc_range_and_freq(ran, sup, tot):
	list1=sorted(ran)
	freq=round((sup/tot)*100,2)
	if list1[0]==list1[-1]:
		return str(round(list1[0],1))+"%", str(int(sup))+"/"+str(int(tot))+" ("+str(round(freq,3))+"%)"
	else:
		return str(round(list1[0],1))+"%-"+str(round(list1[-1],1))+"%", str(int(sup))+"/"+str(int(tot))+" ("+str(round(freq,3))+"%)"		

def calcfreperc(v):
	aa=v.split("/")
	bb=float(aa[0])/float(aa[1])*100
	return str(round(bb,3))+"%"

def make_header(dats):
	header1=['', 'Query', 'Genes', '', '']
	header2=["","","Disrupted","Deleted","Percentil"]
	aa=True
	for el in dats:
		if el!="ClinVar" and el!="ClinGen":
			header1.append(el)
			header1.append('')
			header1.append('')
			header1.append('')
			header1.append('')
		elif (el=="ClinVar" and aa==True and "ClinGen" in dats) or (el=="ClinGen" and aa==True and "ClinVar" in dats):
			header1=header1+['Clin Gen_Var ben.','', '','','', 'Clin Gen_Var likl.ben.','','','','','Clin Gen_Var Uncertain','','','','', 'Clin Gen_Var likl.path','','','','','Clin Gen_Var path','','','','']
			aa=False
		elif el=="ClinGen" and aa==True and "ClinVar" not in dats:
			header1=header1+['Clin Gen ben.', '','','','', 'Clin Gen likl.ben.','','','','','Clin Gen Uncertain','','','','', 'Clin Gen likl.path','','','','','Clin Gen path','','','','']
			aa=False
		elif el=="ClinVar" and aa==True and "ClinGen" not in dats:
			header1=header1+['Clin Var ben.', '','','','', 'Clin Var likl.ben.','','','','','Clin Var Uncertain','','','','', 'Clin Var likl.path','','','','','Clin Var path','','','','']
			aa=False
	gg=5
	while gg<len(header1):
		header2=header2+["#Hits","Type of alteration","Overlap range","Frenquency", "Best hit ID"]
		gg+=5
	return header1, header2


def write_summary(nomenclature_flag, first_flag, stats, ranges, cc, clus, tt, perc, gen, percgene, dats, fgene, is_bp, rr, ovl):#falta o dats
	brek_ranges=[[1,10],[10,20],[20,30],[30,40],[40,50],[50,60],[60,70],[70,80],[80,90],[90,100]]
	best_hit=test_best_hit.perform_best_hit(rr)####
	if first_flag==True:
		wssumary.append(['', 'Variant overlap summary'])
		wssumary.append([])
		wssumary.append(['', 'Inputted parameters'])
		wssumary.append(['', 'Date', dat])
		aa=""
		for el in dats:
			aa+=(el+",")
		wssumary.append(['', 'Genome Version', gen])
		wssumary.append(['', 'Alteration Data', tt])
		wssumary.append(['', 'Alteration Database', aa[:-1]])
		if ovl=="mutual":
			wssumary.append(['', 'Mutual overlap with cutoff percentage', perc])
		if ovl!="mutual":
			wssumary.append(['', 'Query comprised by the reference'])
		if is_bp==True:
			wssumary.append(['', 'Breakpoint flanking region', fgene])###############
		wssumary.append(['', 'Flanking region size for genes', percgene])
		if is_bp==False:
			wssumary.append([])
		wssumary.append([])
		wssumary.append(['', 'Overlap results summary'])
		header1, header2=make_header(dats)
		wssumary.append(header1)
		wssumary.append(header2)
		for key,value in stats.items():
			a=0
			if "_" in key:
				while a<len(value):
					if a==0:
						aa=["",key, clus[key][-2], clus[key][-1], str(brek_ranges[a][0])+"-"+str(brek_ranges[a][1])]
						hiter=best_hit[key]
					if a!=0:
						aa=["","","","", str(brek_ranges[a][0])+"-"+str(brek_ranges[a][1])]
						hiter=""
					if len(value[a])==0:
						aa+=["0"," "," ", " ",hiter]
					if len(value[a])==5:
						aa+=["1",value[a][-1],str(round(float(value[a][1]),1))+"%",value[a][2]+" ("+calcfreperc(value[a][2])+")", hiter]
					if len(value[a])>5:
						aa+=complement_stats(value[a], nomenclature_flag)
						aa+=[hiter]
					wssumary.append(aa)
					aa=[]
					a+=1
			else:
				while a<len(value):
					if a==0:
						aa=["",key, clus[key][-2], clus[key][-1], str(ranges[a][0])+"-"+str(ranges[a][1])]
						hiter=best_hit[key]
					if a!=0:
						aa=["","","","", str(ranges[a][0])+"-"+str(ranges[a][1])]
						hiter=""
					if len(value[a])==0:
						aa+=["0"," "," ", " ", hiter]
					if len(value[a])==5:
						aa+=["1",value[a][-1], str(round(float(value[a][1]),1))+"%",value[a][2]+" ("+calcfreperc(value[a][2])+")", hiter]
					if len(value[a])>5:
						aa+=complement_stats(value[a], nomenclature_flag)
						aa+=[hiter]
					wssumary.append(aa)
					aa=[]
					a+=1
	else:
		bb=15#numero da linha em que comeca
		for key,value in stats.items():
			a=0
			while a<len(value):
				if a==0:
					hiter=best_hit[key]
				if a!=0:
					hiter=""
				if len(value[a])==0:
					wssumary.cell(row=bb, column=cc).value = "0"
					wssumary.cell(row=bb, column=cc+1).value = " "
					wssumary.cell(row=bb, column=cc+2).value = " "
					wssumary.cell(row=bb, column=cc+3).value = " "
					wssumary.cell(row=bb, column=cc+4).value = hiter
				if len(value[a])==5:
					wssumary.cell(row=bb, column=cc).value = "1"
					wssumary.cell(row=bb, column=cc+1).value= value[a][-1]
					wssumary.cell(row=bb, column=cc+2).value = str(round(float(value[a][1]),1))+"%"
					wssumary.cell(row=bb, column=cc+3).value = value[a][2]+" ("+calcfreperc(value[a][2])+")"
					wssumary.cell(row=bb, column=cc+4).value = hiter
				if len(value[a])>5:
					els=complement_stats(value[a], nomenclature_flag)
					wssumary.cell(row=bb, column=cc).value = els[0]
					wssumary.cell(row=bb, column=cc+1).value = els[1]
					wssumary.cell(row=bb, column=cc+2).value = els[2]
					wssumary.cell(row=bb, column=cc+3).value = els[3]
					wssumary.cell(row=bb, column=cc+4).value = hiter
				a+=1
				bb+=1
		cc+=5
	return cc

def check_if_in_output(in_output, clusters, v):
	l=set(clusters.keys())
	ffs=l-in_output
	for el in ffs:
		v[el]=[]
	return v

def write(variants, outfile, flagcoe):
	outfile.append([])
	if flagcoe==True:
		outfile.append(["","Query", "Query Size (bp)","Ref. CNV ID","Ref. CNV size", "Ref. CNV region","Ref. CNV Type of alteration","Ref. CNV Support", "PMID references","", "Overlap % - Query", "Overlap % - Ref. CNV", "Genes and associated phenotype"])
	else:
		outfile.append(["","Query", "Query Size (bp)","Ref. CNV ID","Ref. CNV size", "Ref. CNV region","Ref. CNV Type of alteration","Ref. CNV Support", "PMID references","", "Overlap % - Query", "Overlap % - Ref. CNV"])
	for key, value in variants.items():
		if len(value)>0:
			for el in value:
				tt=[el[0]]#.rsplit("_",1)
				if len(tt)>1:
					gg=tt+el[1:]
				else:
					gg=["",tt[0]]+el[1:]
				outfile.append(gg)
		else:
			tt=[key]#.rsplit("_",1)
			if len(tt)>1:
				tt.append("No hit inside the defined overlap. Please, try a lower overlap cutoff")
				outfile.append(tt)
			else:
				outfile.append(["",key, "No hit inside the defined overlap. Please, try a lower overlap cutoff"])

def calc_merge_best_hit(table, ranges,merge_hits):
	to_use=(table.max_column-5)/5
	merge_use=merge_hits[:to_use]#colunas que vao ter o merge
	nvezes=(table.max_row-14)/len(ranges)
	aa=0
	g=15
	exiting_l=[]
	sub=[]
	while aa<nvezes:
		for el in merge_use:
			exiting_l.append(el+str(g)+":"+el+str(g+len(ranges)-1))
			sub.append(el+str(g))
		g+=len(ranges)
		aa+=1
	return exiting_l, sub



def get_cols(table,ranges):
	cols=["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS"]
	pairs=["B13:B14","C13:D13","F13:J13","K13:O13","P13:T13","U13:Y13","Z13:AD13","AE13:AI13","AJ13:AN13","AO13:AS13"]  
	le=["C","E", "F", "K","P","U","Z","AE","AJ", "AO"]
	merge_hits=["J","O","T","Y","AD","AI","AN","AS"]#####################HITS
	exiting_l, sub=calc_merge_best_hit(table,ranges,merge_hits)
	range_bold=["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK", "AL","AM","AN","AO","AP","AQ","AR","AS"]
	return cols[:table.max_column-1], pairs[:((table.max_column)-5)/5+2]+exiting_l, le[:((table.max_column)-5)/5+2], range_bold[:table.max_column-1], sub#passei o /4 ipara /5

def format_summary(table, ranges, is_bp):
	jj=str(ranges[0][0])
	bot_border=Border(bottom=Side(style='thin'))
	both_border=Border(left=Side(style='thin'),bottom=Side(style="thin"))
	all_border=Border(left=Side(style='thin'), bottom=Side(style="thin"), right=Side(style='thin'), top=Side(style="thin"))
	cols,pairs,le,range_bold, besth=get_cols(table, ranges)
	i=15
	table["B1"].font=Font(bold=True, color="3e3e49", size=14)
	table["B3"].font=Font(bold=True, size=12)
	table["B4"].border=Border(left=Side(style='thick'),top=Side(style="thick"))
	table["C4"].border=Border(right=Side(style='thick'),top=Side(style="thick"))
	table["B5"].border=Border(left=Side(style='thick'))
	table["B6"].border=Border(left=Side(style='thick'))
	table["B7"].border=Border(left=Side(style='thick'))
	table["B8"].border=Border(left=Side(style='thick'))
	table["C5"].border=Border(right=Side(style='thick'))
	table["C6"].border=Border(right=Side(style='thick'))
	table["C7"].border=Border(right=Side(style='thick'))
	table["C8"].border=Border(right=Side(style='thick'))
	if is_bp==True:
		table["C10"].border=Border(right=Side(style='thick'),bottom=Side(style="thick"))
		table["B10"].border=Border(left=Side(style='thick'),bottom=Side(style="thick"))
		table["C9"].border=Border(right=Side(style='thick'))
		table["B9"].border=Border(left=Side(style='thick'))
	else:
		table["C9"].border=Border(right=Side(style='thick'),bottom=Side(style="thick"))
		table["B9"].border=Border(left=Side(style='thick'),bottom=Side(style="thick"))
	table["C8"].border=Border(right=Side(style='thick'))
	table["B12"].border=Border(left=Side(style='thick'),bottom=Side(style="thick"))
	table["C12"].border=Border(right=Side(style='thick'),bottom=Side(style="thick"))
	table["B12"].font=Font(bold=True, size=12)
	table["B12"].border=Border(bottom=Side(style='thick'))
	aa=table.max_column
	table[openpyxl.utils.get_column_letter(aa+1)+"13"].border=Border(left=Side(style='thick'))
	for elh in besth:
		table[elh].font=Font(underline="single", color="0000CC")
	for el in pairs:
		table.merge_cells(el)
	for eli in range_bold:
		table[eli+"13"].font=Font(bold=True, size=12)
		table[eli+"14"].font=Font(bold=True, size=12)
		if eli=="B":
			table[eli+"14"].border=Border(bottom=Side(style="thick"), left=Side(style="thick"), right=Side(style="thin"))
			table[eli+"13"].border=Border(right=Side(style="thin"), left=Side(style="thick"))
		elif eli==openpyxl.utils.get_column_letter(aa):
			table[eli+"14"].border=Border(bottom=Side(style="thick"), right=Side(style="thick"))
			table[eli+"13"].border=Border(right=Side(style="thick"),top=Side(style="thick"))	
		elif eli in le:
			table[eli+"13"].border=Border(top=Side(style="thick"), left=Side(style="thin"))
			table[eli+"14"].border=Border(bottom=Side(style="thick"), left=Side(style="thin"))
		else:	
			table[eli+"13"].border=Border(top=Side(style="thick"))
			table[eli+"14"].border=Border(bottom=Side(style="thick"))		
	while i<=table.max_row:
		gg=table["E"+str(i)]
		bb=gg.value
		if bb!=None:
			hhh=bb.encode("UTF-8")
			if hhh.startswith(jj) or hhh.startswith("1-10"):
				ss=table["C"+str(i)]
				ss.font=Font(italic=True)
				ss=table["D"+str(i)]
				ss.font=Font(italic=True)
				ss=table["B"+str(i)]
				if ss.value!=None:
					hih=ss.value
					if "chr" in hih.encode("UTF-8") or ")" in hih.encode("UTF-8"):
						ss.font=Font(bold=True)
					else:
						ss.font=Font(bold=True, italic=True)
				ss=table["A"+str(i)]
				ss.font=Font(bold=True)
			if hhh.endswith("-100"):
				ss=table["A"+str(i)]
				ss.border=Border(right=Side(style='thick'))
				ss=table[openpyxl.utils.get_column_letter(aa+1)+str(i)]
				ss.border=Border(left=Side(style='thick'))
				for el in cols:
					ss=table[el+str(i)]
					if el not in le:
						ss.border=bot_border
					else:
						ss.border=both_border
			else:
				ss=table["A"+str(i)]
				ss.border=Border(right=Side(style='thick'))
				ss=table[openpyxl.utils.get_column_letter(aa+1)+str(i)]
				ss.border=Border(left=Side(style='thick'))
				for pp in le:
					kk=table[pp+str(i)]
					kk.border=Border(left=Side(style='thin'))
		if i==table.max_row:
			ss=table["A"+str(i)]
			ss.border=Border(right=Side(style='thick'))	
			ss=table[openpyxl.utils.get_column_letter(aa+1)+str(i)]
			ss.border=Border(left=Side(style='thick'))
			for el in cols:
				ss=table[el+str(i)]
				if el not in le:
					ss.border=Border(bottom=Side(style="thick"))
				else:
					ss.border=Border(bottom=Side(style="thick"), left=Side(style="thin"))
		i+=1
	table["B15"].border=Border(top=Side(style='thick'), left=Side(style="thick"))


def format_tables(table, flagcoe):
	if flagcoe==True:
		lets=["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
	else:
		lets=["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
	thick_border=Border(bottom=Side(style='thick'),top=Side(style='thick'))
	thin_border=Border(bottom=Side(style='thin'))
	ot=["D", "I", "J"]
	i=1
	while i<=table.max_row:
		gg=table["B"+str(i)]
		hhh=gg.value
		if hhh!=None:
			bb=hhh.encode("UTF-8")
			if "overlap" in bb:
				table["B"+str(i)].font=Font(bold=True, color="3e3e49", size=14)
				table["C"+str(i)].font=Font(underline="single", color="0000CC", bold=True, size=14)
				if i!=1:
					for el in lets:
						jkj=table[el+str(i-3)]
						if el!="B" and el!=lets[-1]:
							jkj.border=Border(bottom=Side(style='thick'))
						elif el=="B":
							jkj.border=Border(bottom=Side(style='thick'), left=Side(style="thick"))
						elif el==lets[-1]:
							jkj.border=Border(bottom=Side(style='thick'), right=Side(style="thick"))
			elif bb.startswith("Query"):
				for el in lets:
					aa=table[el+str(i)]
					aa.font=Font(bold=True)
					if el!="B" and el!=lets[-1]:
						aa.border=thick_border
					elif el=="B":
						aa.border=Border(bottom=Side(style='thick'),top=Side(style='thick'), left=Side(style="thick"))
					elif el==lets[-1]:
						aa.border=Border(bottom=Side(style='thick'),top=Side(style='thick'), right=Side(style="thick"))
			elif "website" in bb:
				table["B"+str(i)].font=Font(underline="single", color="0000CC", bold=True, size=14)
				table["C"+str(i)].font=Font(bold=True, size=14)
			elif "Date" in bb:
				table["B"+str(i)].font=Font(bold=True, size=14)
			else:
				if table["D"+str(i)].value==None:
					for el in lets:
						jkj=table[el+str(i)]
						jkj.font=Font(color="3e3e49")
						if el!="B" and el!=lets[-1]:
							jkj.border=thin_border
						elif el=="B":
							ggg=jkj.value
							if "chr" not in ggg.encode("UTF-8") and ")" not in ggg.encode("UTF-8"):
								jkj.font=Font(italic=True, color="3e3e49")
							jkj.border=Border(bottom=Side(style='thin'), left=Side(style="thick"))
						elif el==lets[-1]:
							jkj.border=Border(bottom=Side(style='thin'), right=Side(style="thick"))
				else:
					ggg=table["B"+str(i)].value
					if "chr" not in ggg.encode("UTF-8") and ")" not in ggg.encode("UTF-8"):
						table["B"+str(i)].font=Font(italic=True)
					if table["B"+str(i)].value!=table["B"+str(i+1)].value:
						for el in lets:
							jkj=table[el+str(i)]
							if el!="B" and el!=lets[-1]:
								jkj.border=thin_border
							elif el=="B":
								jkj.border=Border(bottom=Side(style='thin'), left=Side(style="thick"))
							elif el==lets[-1]:
								jkj.border=Border(bottom=Side(style='thin'), right=Side(style="thick"))
					elif table["B"+str(i)].value==table["B"+str(i+1)].value:
						jkj=table["B"+str(i)]
						jkj.border=Border(left=Side(style="thick"))
						jkj=table["M"+str(i)]
						if flagcoe==True:
							jkj.border=Border(right=Side(style="thick"))	
						else:
							jkj.border=Border(left=Side(style="thick"))					
					for hh in ot:
						n=table[hh+str(i)]
						n.font=Font(underline="single", color="0000CC")
		if i==table.max_row:
			jk=table["B"+str(i)].value
			if jk!=None:
				aaa=i
			else:
				aaa=i-1
			for el in lets:
				jkj=table[el+str(aaa)]
				if el!="B" and el!=lets[-1]:
					jkj.border=Border(bottom=Side(style='thick'))
				elif el=="B":
					jkj.border=Border(bottom=Side(style='thick'), left=Side(style="thick"))
				elif el==lets[-1]:
					jkj.border=Border(bottom=Side(style='thick'), right=Side(style="thick"))
		i+=1
		
def exect(form):
	ovl=form["ovl"].value#type of overlap
	dat= time.strftime("%d-%m-%Y") 
	global dat
	version=form["version"].value#versao
	try:
		perc=form["perc"].value#overlap cutoff
	except KeyError:
		perc="70"
	try:
		percgene=form["percgene"].value#flaking egion gene
	except KeyError:
		percgene="0bp"
	try:
		bpflank=form["bpflank"].value# franking breakpoint
	except KeyError:
		bpflank="0"
	tt=form["tt"].value#loss gain
	#dats=form["dats[]"].value#dgv, 1000genomes
	dats=[]
	try:
		for i in form["dats[]"]:
			dats.append(i.value)
	except TypeError:
		dats.append(form["dats[]"].value)
	#hg19 dic for data
	datahg19loss={'DGV':'DGV_loss_hg19', '1000Genomes':'1000_Genomes_loss_hg19', 'ClinGenben':'clingen_ben_loss_hg19', 'ClinVarben':'clinVar_loss_Benign_hg19', 'CoeCoop':'Coe_Cooper_loss_DD_hg19', 'ClinGenlben':'clingen_like_ben_loss_hg19', 'ClinVarlben':'clinVar_loss_Likely_benign_hg19','ClinGenun':'clingen_uncertain_loss_hg19', 'ClinVarun':'clinVar_loss_uncertain_hg19','ClinGenlpat':'clingen_like_path_loss_hg19', 'ClinVarlpat':'clinVar_loss_Likely_path_hg19', 'ClinGenpat':'clinGen_path_loss_hg19', 'ClinVarpat':'clinVar_loss_Path_hg19'}
	datahg19gain={'DGV':'DGV_gain_hg19', '1000Genomes':'1000_Genomes_gain_hg19', 'ClinGenben':'clingen_ben_gain_hg19', 'ClinVarben':'clinVar_gain_Benign_hg19', 'CoeCoop':'Coe_Cooper_Gain_DD_hg19', 'ClinGenlben':'clingen_like_ben_gain_hg19', 'ClinVarlben':'clinVar_gain_Likely_benign_hg19','ClinVarun':'clinVar_gain_uncertain_hg19', 'ClinGenun':'clingen_uncertain_gain_hg19','ClinGenlpat':'clingen_like_path_gain_hg19', 'ClinVarlpat':'clinVar_gain_Likely_path_hg19', 'ClinGenpat':'clinGen_path_gain_hg19', 'ClinVarpat':'clinVar_gain_Path_hg19'}
	datahg19inversion={'DGV':'DGV_inversions_hg19', '1000Genomes':'1000_Genomes_inversions_hg19'}
	datahg19all={'DGV':'DGV_all_hg19', '1000Genomes':'1000_Genomes_all_hg19', 'ClinGenben':'clingen_ben_all_hg19', 'ClinVarben':'clinVar_all_Benign_hg19', 'CoeCoop':'Coe_Cooper_all_DD_hg19', 'ClinGenlben':'clingen_like_ben_all_hg19', 'ClinVarlben':'clinVar_all_Likely_Benign_hg19','ClinGenun':'clingen_uncertain_all_hg19', 'ClinVarun':'clinVar_all_uncertain_hg19','ClinGenlpat':'clingen_like_path_all_hg19', 'ClinVarlpat':'clinVar_all_Likely_path_hg19', 'ClinGenpat':'clinGen_path_all_hg19', 'ClinVarpat':'clinVar_all_Path_hg19'}
	#hg38 dic for data
	datahg38loss={'DGV':'DGV_loss_hg38', '1000Genomes':'1000_Genomes_loss_hg38', 'ClinGenben':'clingen_benign_loss_hg38', 'ClinVarben':'clinVar_loss_Benign_hg38', 'CoeCoop':'Coe_Cooper_loss_DD_hg38', 'ClinGenlben':'clingen_likely_ben_loss_hg38', 'ClinVarlben':'clinVar_loss_Likely_Benign_hg38','ClinGenun':'clingen_uncertain_loss_hg38', 'ClinVarun':'clinVar_loss_uncertain_hg38','ClinGenlpat':'clinGen_like_path_loss_hg38', 'ClinVarlpat':'clinVar_loss_Likely_patogenic_hg38', 'ClinGenpat':'clinGen_path_loss_hg38', 'ClinVarpat':'clinVar_loss_patogenic_hg38'}
	datahg38gain={'DGV':'DGV_gain_hg38', '1000Genomes':'1000_Genomes_gain_hg38', 'ClinGenben':'clingen_benign_gain_hg38', 'ClinVarben':'clinVar_gain_Benign_hg38', 'CoeCoop':'Coe_Cooper_Gain_DD_hg38', 'ClinVarlben':'clinVar_gain_Likely_Benign_hg38', 'ClinGenlben':'clingen_likely_ben_gain_hg38','ClinGenun':'clingen_uncertain_gain_hg38', 'ClinVarun':'clinVar_gain_uncertain_hg38','ClinGenlpat':'clinGen_like_path_gain_hg38', 'ClinVarlpat':'clinVar_gain_Likely_patogenic_hg38', 'ClinGenpat':'clinGen_path_gain_hg38', 'ClinVarpat':'clinVar_gain_patogenic_hg38'}
	datahg38inversion={'DGV':'DGV_inversions_hg38', '1000Genomes':'1000_Genomes_inversions_hg38'}
	datahg38all={'DGV':'DGV_all_hg38', '1000Genomes':'1000_Genomes_all_hg38', 'ClinGenben':'clinGen_benign_all_hg38', 'ClinVarben':'clinVar_all_Benign_hg38', 'CoeCoop':'Coe_Cooper_all_DD_hg38', 'ClinGenlben':'clingen_likely_ben_all_hg38', 'ClinVarlben':'clinVar_all_Likely_Benign_hg38','ClinGenun':'clingen_uncertain_all_hg38', 'ClinVarun':'clinVar_all_uncertain_hg38','ClinGenlpat':'clingen_like_path_all_hg38', 'ClinVarlpat':'clinVar_all_Likely_patogenic_hg38', 'ClinGenpat':'clinGen_path_all_hg38', 'ClinVarpat':'clinVar_all_patogenic_hg38'}
	titles={'DGV':["","DGV Database overlap Results", '', '=HYPERLINK("http://dgv.tcag.ca/dgv/app/home","DGV database website")'], '1000Genomes':["","1000 Genomes Database overlap Results",'', '=HYPERLINK("http://www.internationalgenome.org/","1000 genomes website")'], 'ClinGenben':["","ClinGen Benign alterations Database overlap Results",'',  '=HYPERLINK("https://www.clinicalgenome.org/","ClinGen website")'], 'ClinVarben':["","ClinVar Benign alterations Database overlap Results", '', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/clinvar/","ClinVar website")'], 'CoeCoop':["","Developmental Delay Database overlap Results - Case (Coe et al. 2014; Cooper et al. 2011)",''], 'ClinGenlben':["","ClinGen Likelly Benign alterations Database overlap Results", '', '=HYPERLINK("https://www.clinicalgenome.org/","ClinGen website")'], 'ClinVarlben':["","ClinVar Likely Benign alterations Database overlap Results", '', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/clinvar/","ClinVar website")'],'ClinGenun':["","ClinGen uncertain alterations Database overlap Results", '', '=HYPERLINK("https://www.clinicalgenome.org/","ClinGen website")'], 'ClinVarun':["","ClinVar uncertain alterations Database overlap Results", '', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/clinvar/","ClinVar website")'],'ClinGenlpat':["","ClinGen Likely pathogenic alterations Database overlap Results", '', '=HYPERLINK("https://www.clinicalgenome.org/","ClinGen website")'], 'ClinVarlpat':["","ClinVar Likely pathogenic alterations Database overlap Results", '', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/clinvar/","ClinVar website")'], 'ClinGenpat':["","ClinGen Pathogenic alterations Database overlap Results", '', '=HYPERLINK("https://www.clinicalgenome.org/","ClinGen website")'], 'ClinVarpat':["","ClinVar Pathogenic alterations Database overlap Results", '', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/clinvar/","ClinVar website")']}
	sufs=["ClinGenben", "ClinVarben","ClinGenlben","ClinVarlben", "ClinGenun","ClinVarun", "ClinGenlpat",  "ClinVarlpat", "ClinGenpat", "ClinVarpat"]
	sufs2=["ClinGenben", "ClinGenlben", "ClinGenun","ClinGenlpat","ClinGenpat"]
	sufs3=["ClinVarben","ClinVarlben","ClinVarun","ClinVarlpat","ClinVarpat"]
	ttt=form["ttt"].value
	fiile_bp=[]
	fiile=[]
	if ttt=="breakpoint" or ttt=="bpcnv":
		fiile_bp=form["message"].value
	if ttt=="cnv" or ttt=="bpcnv":
		fiile=form["msg"].value
	###########alterar para por as novas opçoes
	print('<aaa><center><font size="5em;"><b>Input parameters:</b></font></center></aaa>')
	print("<aaa><center><b>Genome version:</b>"+version+"</center></aaa>")
	print("<aaa><center><b>Alteration Data:</b>"+tt+"</center></aaa>")
	print("<aaa><center><b>Alteration Database:</b>"+",".join(dats)+"</center></aaa>")
	if ovl=="mutual":
		print("<aaa><center><b>Type of overlap search:</b> Mutual overlap</center></aaa>")
		print("<aaa><center><b>Overlap cutoff percentage:</b>"+perc+"%</center></aaa>")
	else:
		print("<aaa><center><b>Type of overlap search:</b> Query comprised by the reference</center></aaa>")
	print("<aaa><center><b>Breakpoint flanking region:</b>"+bpflank+"</center></aaa>")
	if len(fiile_bp)>1:
		print("<aaa><center><b>Input breakpoint positions:</b></center></aaa>")
		for ee in fiile_bp.split("\n"):
			if len(ee)>5:
				print("<aaa><center>"+ee+"</center></aaa>")
	print("<aaa><center><b>Flanquing region size:</b>"+percgene+"</center></aaa>")
	if len(fiile)>1:
		print("<aaa><center><b>Input CNV or genomic region:</b></center></aaa>")
		for ee in fiile.split("\n"):
			if len(ee)>5:
				print("<aaa><center>"+ee+"</center></aaa>")
	print('<aaa><center><b><font size="5em;">Output:</font></b></center></aaa>')
	if version=="hg38":
		if len(fiile)>1:
			cc=11
			gt=11
			at1=['start_position', 'end_position','external_gene_name', 'gene_biotype']
			global at1
			dic, chrrs=read_bed(fiile, version,percgene)
			name="Database_search_CNV_"+dat+".xlsx"
			outfile=("/usr/local/apache/htdocs/"+name)
			wb=Workbook()
			global wssumary
			check=True
			wssumary=wb.active################
			wssumary.title="Summary"###########################
			for datas in dats:
				if datas!="ClinGen" and datas!="ClinVar":
					ws1=wb.create_sheet()
					ws1.title=datas
					if tt=="loss":
						dic_ch, dic_name, date=read_bd(datahg38loss[datas], chrrs)
					elif tt=="gain":
						dic_ch, dic_name, date=read_bd(datahg38gain[datas], chrrs)
					elif tt=="inversion" and "inversion" in datahg38inversion:
						dic_ch, dic_name, date=read_bd(datahg38inversion[datas], chrrs)
					elif tt=="all":
						dic_ch, dic_name, date=read_bd(datahg38all[datas], chrrs)
					titles[datas].append(date)
					ws1.append(titles[datas][:2])
					ws1.append(titles[datas][2:])
					ws1.append([])
					stats=OrderedDict()
					v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, False, ovl)
					if datas=="CoeCoop":
						write(v, ws1, True)
						format_tables(ws1, True)
					else:
						write(v, ws1, False)
						format_tables(ws1, False)
					if datas==dats[0]:
						cc=write_summary( True, True, stats, ranges, cc, dic, tt, perc, version, percgene, dats,bpflank, False, v, ovl)
					elif datas!=dats[0]:
						cc=write_summary( False, False, stats, ranges, cc, dic, tt, perc, version, percgene, dats,bpflank, False, v, ovl)
					stats=OrderedDict()
				elif check==True:
					if "ClinGen" in dats and "ClinVar" in dats:
						s=sufs
					elif "ClinGen" in dats and "ClinVar" not in dats:
						s=sufs2
					elif "ClinGen" not in dats and "ClinVar" in dats:
						s=sufs3
					for dd in s:
						if dd.startswith("ClinVar") and "ClinGen" in dats:
							ws1.append([])
							ws1.append([])
						else:
							ws1=wb.create_sheet()
							ws1.title=dd
						if tt=="loss":
							dic_ch, dic_name, date=read_bd(datahg38loss[dd], chrrs)
						elif tt=="gain":
							dic_ch, dic_name, date=read_bd(datahg38gain[dd], chrrs)
						elif tt=="inversion" and "inversion" in datahg38inversion:
							dic_ch, dic_name, date=read_bd(datahg38inversion[dd], chrrs)
						elif tt=="all":
							dic_ch, dic_name, date=read_bd(datahg38all[dd], chrrs)
						titles[dd].append(date)
						ws1.append(titles[dd][:2])
						ws1.append(titles[dd][2:])
						#ws1.append(titles[dd])
						ws1.append([])
						v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats,False, ovl)
						write(v, ws1, False)
						if (dd.startswith("ClinGen") and "ClinVar" not in dats) or (dd.startswith("ClinVar")) :
							format_tables(ws1, False)
						if ("ClinVar" in dats and "ClinGen" in dats and s.index(dd)%2==1) or "ClinVar" not in dats or "ClinGen" not in dats:
							if datas==dats[0] and dd==s[0]:
								cc=write_summary( True, True, stats, ranges, cc, dic, tt, perc, version, percgene, dats, bpflank, False, v, ovl)
								stats=OrderedDict()
							else:
								cc=write_summary( False, False, stats, ranges, cc, dic, tt, perc, version, percgene, dats, bpflank, False, v, ovl)
								stats=OrderedDict()
						check=False
			format_summary(wssumary, ranges, False)
			wb.save(filename=outfile)
			print("<aaa><center><a href=/"+name+">Download CNV table!</a></center></aaa>")
		if len(fiile_bp)>1:
			cc=11
			gt=11
			at1=['start_position', 'end_position','external_gene_name', 'gene_biotype']
			global at1
			dic, chrrs=read_bps(fiile_bp, version,percgene, bpflank)
			name2="Database_search_BP_"+dat+".xlsx"
			outfile2=("/usr/local/apache/htdocs/"+name2)
			wb=Workbook()
			global wssumary
			check=True
			wssumary=wb.active################
			wssumary.title="Summary"###########################
			for datas in dats:
				if datas!="ClinGen" and datas!="ClinVar":
					ws1=wb.create_sheet()
					ws1.title=datas
					if tt=="loss":
						dic_ch, dic_name, date=read_bd(datahg38loss[datas], chrrs)
					elif tt=="gain":
						dic_ch, dic_name, date=read_bd(datahg38gain[datas], chrrs)
					elif tt=="inversion" and "inversion" in datahg38inversion:
						dic_ch, dic_name, date=read_bd(datahg38inversion[datas], chrrs)
					elif tt=="all":
						dic_ch, dic_name, date=read_bd(datahg38all[datas], chrrs)
					titles[datas].append(date)
					#ws1.append(titles[datas])
					ws1.append(titles[datas][:2])
					ws1.append(titles[datas][2:])
					ws1.append([])
					stats=OrderedDict()
					v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, True, ovl)
					if datas=="CoeCoop":
						write(v, ws1, True)
						format_tables(ws1, True)
					else:
						write(v, ws1, False)
						format_tables(ws1, False)
					if datas==dats[0]:
						gt=write_summary( True, True, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
					elif datas!=dats[0]:
						gt=write_summary( False, False, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
					stats=OrderedDict()
				elif check==True:
					if "ClinGen" in dats and "ClinVar" in dats:
						s=sufs
					elif "ClinGen" in dats and "ClinVar" not in dats:
						s=sufs2
					elif "ClinGen" not in dats and "ClinVar" in dats:
						s=sufs3
					for dd in s:
						if dd.startswith("ClinVar") and "ClinGen" in dats:
							ws1.append([])
							ws1.append([])
						else:
							ws1=wb.create_sheet()
							ws1.title=dd
						if tt=="loss":
							dic_ch, dic_name, date=read_bd(datahg38loss[dd], chrrs)
						elif tt=="gain":
							dic_ch, dic_name, date=read_bd(datahg38gain[dd], chrrs)
						elif tt=="inversion" and "inversion" in datahg38inversion:
							dic_ch, dic_name, date=read_bd(datahg38inversion[dd], chrrs)
						elif tt=="all":
							dic_ch, dic_name, date=read_bd(datahg38all[dd], chrrs)
						titles[dd].append(date)
						ws1.append(titles[dd][:2])
						ws1.append(titles[dd][2:])
						#ws1.append(titles[dd])
						ws1.append([])
						v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, True, ovl)
						write(v, ws1, False)
						if (dd.startswith("ClinGen") and "ClinVar" not in dats) or (dd.startswith("ClinVar")) :
							format_tables(ws1, False)
						if ("ClinVar" in dats and "ClinGen" in dats and s.index(dd)%2==1) or "ClinVar" not in dats or "ClinGen" not in dats:
							if datas==dats[0] and dd==s[0]:
								gt=write_summary( True, True, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
								stats=OrderedDict()
							else:
								gt=write_summary( False, False, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
								stats=OrderedDict()
						check=False
			format_summary(wssumary, ranges, True)
			wb.save(filename=outfile2)
			print("<aaa><center><a href=/"+name2+">Download breakpoint table!</a></center></aaa>")
	if version=="hg19":
		at1=['start_position', 'end_position','external_gene_id', 'gene_biotype']
		global at1
		if len(fiile)>1:
			cc=11
			gt=11
			dic, chrrs=read_bed(fiile, version,percgene)
			name="Database_search_CNV_"+dat+".xlsx"
			outfile=("/usr/local/apache/htdocs/"+name)
			wb=Workbook()
			global wssumary
			check=True
			wssumary=wb.active################
			wssumary.title="Summary"###########################
			for datas in dats:
				if datas!="ClinGen" and datas!="ClinVar":
					ws1=wb.create_sheet()
					ws1.title=datas
					if tt=="loss":
						dic_ch, dic_name, date=read_bd(datahg19loss[datas], chrrs)
					elif tt=="gain":
						dic_ch, dic_name, date=read_bd(datahg19gain[datas], chrrs)
					elif tt=="inversion" and "inversion" in datahg19inversion:
						dic_ch, dic_name, date=read_bd(datahg19inversion[datas], chrrs)
					elif tt=="all":
						dic_ch, dic_name, date=read_bd(datahg19all[datas], chrrs)
					titles[datas].append(date)
					#ws1.append(titles[datas])
					ws1.append(titles[datas][:2])
					ws1.append(titles[datas][2:])
					ws1.append([])
					stats=OrderedDict()
					v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, False, ovl)
					if datas=="CoeCoop":
						write(v, ws1, True)
						format_tables(ws1, True)
					else:
						write(v, ws1, False)
						format_tables(ws1, False)
					if datas==dats[0]:
						cc=write_summary( True, True, stats, ranges, cc, dic, tt, perc, version, percgene, dats,bpflank, False, v, ovl)
					elif datas!=dats[0]:
						cc=write_summary( False, False, stats, ranges, cc, dic, tt, perc, version, percgene, dats,bpflank, False, v, ovl)
					stats=OrderedDict()
				elif check==True:
					if "ClinGen" in dats and "ClinVar" in dats:
						s=sufs
					elif "ClinGen" in dats and "ClinVar" not in dats:
						s=sufs2
					elif "ClinGen" not in dats and "ClinVar" in dats:
						s=sufs3
					for dd in s:
						if dd.startswith("ClinVar") and "ClinGen" in dats:
							ws1.append([])
							ws1.append([])
						else:
							ws1=wb.create_sheet()
							ws1.title=dd
						if tt=="loss":
							dic_ch, dic_name, date=read_bd(datahg19loss[dd], chrrs)
						elif tt=="gain":
							dic_ch, dic_name, date=read_bd(datahg19gain[dd], chrrs)
						elif tt=="inversion" and "inversion" in datahg19inversion:
							dic_ch, dic_name, date=read_bd(datahg19inversion[dd], chrrs)
						elif tt=="all":
							dic_ch, dic_name, date=read_bd(datahg19all[dd], chrrs)
						titles[dd].append(date)
						#ws1.append(titles[dd])
						ws1.append(titles[dd][:2])
						ws1.append(titles[dd][2:])
						ws1.append([])
						v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats,False, ovl)
						write(v, ws1, False)
						if (dd.startswith("ClinGen") and "ClinVar" not in dats) or (dd.startswith("ClinVar")) :
							format_tables(ws1, False)
						if ("ClinVar" in dats and "ClinGen" in dats and s.index(dd)%2==1) or "ClinVar" not in dats or "ClinGen" not in dats:
							if datas==dats[0] and dd==s[0]:
								cc=write_summary( True, True, stats, ranges, cc, dic, tt, perc, version, percgene, dats, bpflank, False, v, ovl)
								stats=OrderedDict()
							else:
								cc=write_summary( False, False, stats, ranges, cc, dic, tt, perc, version, percgene, dats, bpflank, False, v, ovl)
								stats=OrderedDict()
						check=False
			format_summary(wssumary, ranges, False)
			wb.save(filename=outfile)
			print("<aaa><center><a href=/"+name+">Download CNV table!</a></center></aaa>")
		if len(fiile_bp)>1:
			cc=11
			gt=11
			at1=['start_position', 'end_position','external_gene_id', 'gene_biotype']
			dic, chrrs=read_bps(fiile_bp, version,percgene, bpflank)
			name2="Database_search_BP_"+dat+".xlsx"
			outfile2=("/usr/local/apache/htdocs/"+name2)
			wb=Workbook()
			global wssumary
			check=True
			wssumary=wb.active################
			wssumary.title="Summary"###########################
			for datas in dats:
				if datas!="ClinGen" and datas!="ClinVar":
					ws1=wb.create_sheet()
					ws1.title=datas
					if tt=="loss":
						dic_ch, dic_name, date=read_bd(datahg19loss[datas], chrrs)
					elif tt=="gain":
						dic_ch, dic_name, date=read_bd(datahg19gain[datas], chrrs)
					elif tt=="inversion" and "inversion" in datahg19inversion:
						dic_ch, dic_name, date=read_bd(datahg19inversion[datas], chrrs)
					elif tt=="all":
						dic_ch, dic_name, date=read_bd(datahg19all[datas], chrrs)
					titles[datas].append(date)
					#ws1.append(titles[datas])
					ws1.append(titles[datas][:2])
					ws1.append(titles[datas][2:])
					ws1.append([])
					stats=OrderedDict()
					v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, True, ovl)
					if datas=="CoeCoop":
						write(v, ws1, True)
						format_tables(ws1, True)
					else:
						write(v, ws1, False)
						format_tables(ws1, False)
					if datas==dats[0]:
						gt=write_summary( True, True, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
					elif datas!=dats[0]:
						gt=write_summary( False, False, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
					stats=OrderedDict()
				elif check==True:
					if "ClinGen" in dats and "ClinVar" in dats:
						s=sufs
					elif "ClinGen" in dats and "ClinVar" not in dats:
						s=sufs2
					elif "ClinGen" not in dats and "ClinVar" in dats:
						s=sufs3
					for dd in s:
						if dd.startswith("ClinVar") and "ClinGen" in dats:
							ws1.append([])
							ws1.append([])
						else:
							ws1=wb.create_sheet()
							ws1.title=dd
						if tt=="loss":
							dic_ch, dic_name, date=read_bd(datahg19loss[dd], chrrs)
						elif tt=="gain":
							dic_ch, dic_name, date=read_bd(datahg19gain[dd], chrrs)
						elif tt=="inversion" and "inversion" in datahg19inversion:
							dic_ch, dic_name, date=read_bd(datahg19inversion[dd], chrrs)
						elif tt=="all":
							dic_ch, dic_name, date=read_bd(datahg19all[dd], chrrs)
						titles[dd].append(date)
						ws1.append(titles[dd][:2])
						ws1.append(titles[dd][2:])
						#ws1.append(titles[dd])
						ws1.append([])
						v, stats, ranges=cross_values(dic, dic_ch, dic_name, perc, stats, True, ovl)
						write(v, ws1, False)
						if (dd.startswith("ClinGen") and "ClinVar" not in dats) or (dd.startswith("ClinVar")) :
							format_tables(ws1, False)
						if ("ClinVar" in dats and "ClinGen" in dats and s.index(dd)%2==1) or "ClinVar" not in dats or "ClinGen" not in dats:
							if datas==dats[0] and dd==s[0]:
								gt=write_summary( True, True, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
								stats=OrderedDict()
							else:
								gt=write_summary( False, False, stats, ranges, gt, dic, tt, perc, version, percgene, dats, bpflank, True, v, ovl)
								stats=OrderedDict()
						check=False
			format_summary(wssumary, ranges, True)
			wb.save(filename=outfile2)
			print("<aaa><center><a href=/"+name2+">Download breakpoint table!</a></center></aaa>")
	print('<aaa><center>If you using this tool please acknowledge either by <i>This table was performed by the CNV-Content tool</i> or by citing <a href="http://www.insa.min-saude.pt/category/areas-de-atuacao/genetica-humana/">our reference publication</a></center><br><br />')
	
def remake():
	print """
<html>

<body>

<center><hgh><button onclick="goBack()">New search</button></center></hgh>

<script>
function goBack() {
  window.location.replace("../CNV-ConTool.py");
}
</script>

</body>
</html>"""
	print('<br><br />')
	print('<b><center><a href=/tadgctV2_tutorial.pdf>Reference manual</a></center></b><br><br />')
	print('<center><address>Correspondance: <a href="mailto:doencasgenomicas@insa.min-saude.pt">Genomic Diseases Group</a>.<br><br /></address>')
	print('<center><aaa><a href="http://www.insa.min-saude.pt/category/areas-de-atuacao/genetica-humana/">Department of Human Genetics</a></aaa></center><br>')
	print('<p>National Institute of Health Doutor Ricardo Jorge</p> </aaa></center>')
	print('<center><img src="https://cld.pt/dl/download/1f715328-21eb-49bd-b04c-b46bf2f08c61/aaa.jpg" width=641 height=122 border=0 alt=""><br><br />')
	print('<center><p><rodape>This file was last modified 10/07/2019</p></font></rodape>')

